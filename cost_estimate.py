#!/usr/bin/env python3
# Copyright 2025 Nicolas Cravino
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""
Generate a Cost Impact Analysis Excel report from actual pipeline run data.

Extrapolates real per-example costs and timings across population × epoch
ranges so a user can pre-estimate before committing to a full run.

Usage (standalone):
    python cost_estimate.py runs/2026-02-24_15-44-54
    python cost_estimate.py runs/2026-02-24_15-44-54 --output my_estimate.xlsx

Usage (from pipeline):
    from cost_estimate import generate_cost_estimate
    generate_cost_estimate(run_dir="runs/...", run_results={...}, cost_info={...})
"""

import argparse
import json
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ───────────────────────────────────────────────────────────
DARK_BG = PatternFill("solid", fgColor="1B2A4A")
HEADER_BG = PatternFill("solid", fgColor="3A5BA0")
LIGHT_BG = PatternFill("solid", fgColor="E8EDF5")
ALT_BG = PatternFill("solid", fgColor="F5F7FB")
WHITE_BG = PatternFill("solid", fgColor="FFFFFF")
GREEN_BG = PatternFill("solid", fgColor="D4EDDA")

WHITE_FONT_LG = Font(name="Arial", color="FFFFFF", bold=True, size=14)
WHITE_FONT_SM = Font(name="Arial", color="FFFFFF", size=10)
HEADER_FONT = Font(name="Arial", color="FFFFFF", bold=True, size=11)
BLUE_FONT = Font(name="Arial", color="0000FF", size=11)
BLACK_FONT = Font(name="Arial", color="000000", size=11)
BLACK_FONT_B = Font(name="Arial", color="000000", bold=True, size=11)
TITLE_FONT = Font(name="Arial", color="1B2A4A", bold=True, size=13)
NOTE_FONT = Font(name="Arial", color="666666", italic=True, size=10)

BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

POPULATIONS = [3, 5, 10, 15, 20, 25, 30, 48]
EPOCHS = [1, 2, 4, 6, 8, 10]

MODEL_PRICING = {
    "claude-haiku-4-5":   ("Claude Haiku 3.5",  0.80,  4.00),
    "claude-sonnet-4-6":  ("Claude Sonnet 4.6",  3.00, 15.00),
    "claude-opus-4-6":    ("Claude Opus 4.6",   15.00, 75.00),
}


def _cell(ws, ref, value, font=BLACK_FONT, fill=None, fmt=None, align=CENTER, border=BORDER):
    c = ws[ref]
    c.value = value
    c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    c.alignment = align
    c.border = border
    return c


def generate_cost_estimate(
    run_dir: str,
    run_results: dict | None = None,
    cost_info: dict | None = None,
    output: str | None = None,
) -> str:
    """Generate cost impact analysis XLSX. Returns output path."""
    run_path = Path(run_dir)

    # Load data from files if not passed directly
    if cost_info is None:
        cost_file = run_path / "cost_report.json"
        if cost_file.exists():
            with open(cost_file) as f:
                cost_info = json.load(f)
        else:
            raise FileNotFoundError(f"No cost_report.json in {run_dir}")

    if run_results is None:
        results_file = run_path / "run_results.json"
        if results_file.exists():
            with open(results_file) as f:
                run_results = json.load(f)
        else:
            run_results = {}

    # Extract actuals
    n_examples = cost_info.get("examples_generated", 1)
    model_name = cost_info.get("model", "claude-sonnet-4-6")
    avg_input_tokens = round(cost_info.get("total_input_tokens", 183) / max(n_examples, 1))
    avg_output_tokens = round(cost_info.get("total_output_tokens", 8192) / max(n_examples, 1))
    cost_per_example = cost_info.get("avg_cost_per_example_usd", 0.1234)
    total_cost = cost_info.get("total_cost_usd", 0)

    # Timing from step_results
    steps = run_results.get("step_results", {})
    api_total_sec = steps.get("generate_teacher_data", {}).get("elapsed_seconds", 0)
    baseline_sec = steps.get("baseline_inference", {}).get("elapsed_seconds", 30)
    train_sec = steps.get("finetune", {}).get("elapsed_seconds", 0)
    distill_sec = steps.get("distilled_inference", {}).get("elapsed_seconds", 225)

    api_sec_per_ex = round(api_total_sec / max(n_examples, 1), 1)
    # iters ≈ examples (batch_size=1, 1 epoch in the test)
    run_epochs = run_results.get("epochs", 1)
    iters_in_run = n_examples * run_epochs
    train_sec_per_iter = round(train_sec / max(iters_in_run, 1), 1)

    run_id = run_results.get("run_id", run_path.name)

    # Pricing for the model used
    pricing_lookup = {
        "claude-haiku-4-5": (0.80, 4.00), "claude-haiku-4-5-20251001": (0.80, 4.00),
        "claude-sonnet-4-6": (3.00, 15.00), "claude-sonnet-4-5": (3.00, 15.00),
        "claude-sonnet-4-5-20250929": (3.00, 15.00), "claude-sonnet-4-20250514": (3.00, 15.00),
        "claude-opus-4-6": (15.00, 75.00), "claude-opus-4-5": (15.00, 75.00),
        "claude-opus-4-5-20251101": (15.00, 75.00),
    }
    input_price, output_price = pricing_lookup.get(model_name, (3.00, 15.00))

    # ── Build workbook ───────────────────────────────────────────────
    wb = Workbook()

    # ── Sheet 1: Cost Estimator ──────────────────────────────────────
    ws = wb.active
    ws.title = "Cost Estimator"
    ws.sheet_properties.tabColor = "1B2A4A"

    for col_letter, w in {"A": 4, "B": 30, "C": 20, "D": 24, "E": 18, "F": 18, "G": 18, "H": 18, "I": 18}.items():
        ws.column_dimensions[col_letter].width = w

    # Title banner
    for c in range(1, 10):
        ws.cell(row=1, column=c).fill = DARK_BG
        ws.cell(row=2, column=c).fill = DARK_BG
        ws.cell(row=3, column=c).fill = DARK_BG
    ws.merge_cells("B1:I1")
    ws.merge_cells("B2:I2")
    _cell(ws, "B1", "Distillation Attack — Cost Impact Analysis", WHITE_FONT_LG, align=Alignment(horizontal="left", vertical="bottom"))
    _cell(ws, "B2", f"Based on actual run: {run_id}  ({n_examples} examples, {run_epochs} epoch{'s' if run_epochs != 1 else ''})", WHITE_FONT_SM, align=Alignment(horizontal="left", vertical="top"))
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    # Assumptions section
    row = 5
    ws.merge_cells(f"B{row}:D{row}")
    _cell(ws, f"B{row}", "Assumptions (edit blue cells to re-estimate)", TITLE_FONT, align=LEFT)

    assumptions = [
        ("Teacher Model", model_name, None, ""),
        ("Input Price ($/MTok)", input_price, "$#,##0.00", "Anthropic pricing"),
        ("Output Price ($/MTok)", output_price, "$#,##0.00", "Anthropic pricing"),
        ("Avg Input Tokens / Request", avg_input_tokens, "#,##0", f"Measured: {cost_info.get('total_input_tokens', 0):,} / {n_examples}"),
        ("Avg Output Tokens / Request", avg_output_tokens, "#,##0", f"Measured: {cost_info.get('total_output_tokens', 0):,} / {n_examples}"),
        ("Student Model", run_results.get("student_model", "Llama 3.1 8B 4-bit"), None, ""),
    ]

    a_start = row + 1
    for i, (label, val, fmt, note) in enumerate(assumptions):
        r = a_start + i
        ws.row_dimensions[r].height = 22
        _cell(ws, f"B{r}", label, BLACK_FONT, LIGHT_BG, align=LEFT)
        _cell(ws, f"C{r}", val, BLUE_FONT, fmt=fmt)
        _cell(ws, f"D{r}", note, NOTE_FONT, align=LEFT, border=Border())

    # Derived: cost per example
    r_cpe = a_start + len(assumptions)
    ws.row_dimensions[r_cpe].height = 22
    _cell(ws, f"B{r_cpe}", "Cost per Example (derived)", BLACK_FONT_B, GREEN_BG, align=LEFT)
    # Formula: (avg_input * input_price + avg_output * output_price) / 1_000_000
    inp_cell = f"C{a_start + 3}"   # avg input tokens
    out_cell = f"C{a_start + 4}"   # avg output tokens
    inp_price_cell = f"C{a_start + 1}"  # input price
    out_price_cell = f"C{a_start + 2}"  # output price
    _cell(ws, f"C{r_cpe}", f"=({inp_cell}*{inp_price_cell}+{out_cell}*{out_price_cell})/1000000",
          Font(name="Arial", color="000000", bold=True, size=11), GREEN_BG, "$#,##0.0000")
    _cell(ws, f"D{r_cpe}", "Per teacher API call", NOTE_FONT, align=LEFT, border=Border())

    # ── API Cost Matrix ──────────────────────────────────────────────
    m_start = r_cpe + 3
    ws.merge_cells(f"B{m_start}:H{m_start}")
    _cell(ws, f"B{m_start}", "Estimated API Cost by Number of Training Examples", TITLE_FONT, align=LEFT)
    ws.row_dimensions[m_start].height = 28

    hdr = m_start + 1
    ws.row_dimensions[hdr].height = 28
    _cell(ws, f"B{hdr}", "Examples", HEADER_FONT, HEADER_BG)
    _cell(ws, f"C{hdr}", "API Cost", HEADER_FONT, HEADER_BG)
    _cell(ws, f"D{hdr}", "Δ vs Test Run", HEADER_FONT, HEADER_BG)

    for i, pop in enumerate(POPULATIONS):
        r = hdr + 1 + i
        ws.row_dimensions[r].height = 24
        bg = ALT_BG if i % 2 == 0 else WHITE_BG
        _cell(ws, f"B{r}", f"{pop} examples", BLACK_FONT_B, bg)
        _cell(ws, f"C{r}", f"={pop}*C{r_cpe}", BLACK_FONT, bg, "$#,##0.00")
        _cell(ws, f"D{r}", f"=C{r}/C{hdr + 1}", BLACK_FONT, bg, '#,##0.0"x"')

    note_r = hdr + 1 + len(POPULATIONS) + 1
    ws.merge_cells(f"B{note_r}:H{note_r}")
    _cell(ws, f"B{note_r}",
          "API cost depends only on the number of training examples. Epochs affect local GPU time, not API spend.",
          NOTE_FONT, align=LEFT, border=Border())

    # ── Time Estimate Matrix ─────────────────────────────────────────
    t_start = note_r + 3
    ws.merge_cells(f"B{t_start}:H{t_start}")
    _cell(ws, f"B{t_start}", "Estimated Total Pipeline Time (minutes)", TITLE_FONT, align=LEFT)
    ws.row_dimensions[t_start].height = 28

    time_assumptions = [
        ("API Time / Example (sec)", api_sec_per_ex, "#,##0.0", f"Measured: {api_total_sec:.0f}s / {n_examples} ex"),
        ("Baseline Inference (sec)", round(baseline_sec, 1), "#,##0.0", "Fixed per run"),
        ("Training Time / Iter (sec)", train_sec_per_iter, "#,##0.0", f"Measured: {train_sec:.0f}s / {iters_in_run} iters"),
        ("Distilled Inference (sec)", round(distill_sec, 1), "#,##0.0", "Fixed per run"),
    ]

    ta_start = t_start + 1
    for i, (label, val, fmt, note) in enumerate(time_assumptions):
        r = ta_start + i
        ws.row_dimensions[r].height = 22
        _cell(ws, f"B{r}", label, BLACK_FONT, LIGHT_BG, align=LEFT)
        _cell(ws, f"C{r}", val, BLUE_FONT, fmt=fmt)
        _cell(ws, f"D{r}", note, NOTE_FONT, align=LEFT, border=Border())

    api_t = f"C{ta_start}"
    base_t = f"C{ta_start + 1}"
    train_t = f"C{ta_start + 2}"
    dist_t = f"C{ta_start + 3}"

    thdr = ta_start + len(time_assumptions) + 1
    ws.row_dimensions[thdr].height = 28
    _cell(ws, f"B{thdr}", "Examples ↓ / Epochs →", HEADER_FONT, HEADER_BG)
    for j, ep in enumerate(EPOCHS):
        col = get_column_letter(3 + j)
        _cell(ws, f"{col}{thdr}", f"{ep} epoch{'s' if ep > 1 else ''}", HEADER_FONT, HEADER_BG)

    for i, pop in enumerate(POPULATIONS):
        r = thdr + 1 + i
        ws.row_dimensions[r].height = 24
        bg = ALT_BG if i % 2 == 0 else WHITE_BG
        _cell(ws, f"B{r}", f"{pop} examples", BLACK_FONT_B, bg)
        for j, ep in enumerate(EPOCHS):
            col = get_column_letter(3 + j)
            # total = (pop * api_time + baseline + pop * ep * train_per_iter + distilled) / 60
            _cell(ws, f"{col}{r}",
                  f"=({pop}*{api_t}+{base_t}+{pop}*{ep}*{train_t}+{dist_t})/60",
                  BLACK_FONT, bg, "#,##0.0")

    tn_r = thdr + 1 + len(POPULATIONS) + 1
    ws.merge_cells(f"B{tn_r}:H{tn_r}")
    _cell(ws, f"B{tn_r}",
          f"Values in minutes. Extrapolated from {n_examples}-example test run on Apple M-series.",
          NOTE_FONT, align=LEFT, border=Border())

    # ── Sheet 2: Model Pricing ───────────────────────────────────────
    ws2 = wb.create_sheet("Model Pricing")
    ws2.sheet_properties.tabColor = "3A5BA0"
    for col_letter, w in {"A": 4, "B": 32, "C": 20, "D": 20, "E": 22, "F": 22}.items():
        ws2.column_dimensions[col_letter].width = w

    for c in range(1, 7):
        ws2.cell(row=1, column=c).fill = DARK_BG
    ws2.merge_cells("B1:E1")
    _cell(ws2, "B1", "Anthropic Model Pricing Reference", WHITE_FONT_LG,
          align=Alignment(horizontal="left", vertical="center"))
    ws2.row_dimensions[1].height = 36

    headers = ["Model", "Input ($/MTok)", "Output ($/MTok)", "Est. Cost / Example", "Est. Cost / 48 Ex"]
    for j, h in enumerate(headers):
        _cell(ws2, ws2.cell(row=3, column=2+j).coordinate, h, HEADER_FONT, HEADER_BG)
    ws2.row_dimensions[3].height = 28

    for i, (key, (name, inp, out)) in enumerate(MODEL_PRICING.items()):
        r = 4 + i
        bg = ALT_BG if i % 2 == 0 else WHITE_BG
        ws2.row_dimensions[r].height = 24
        _cell(ws2, f"B{r}", name, BLACK_FONT_B, bg, align=LEFT)
        _cell(ws2, f"C{r}", inp, BLUE_FONT, bg, "$#,##0.00")
        _cell(ws2, f"D{r}", out, BLUE_FONT, bg, "$#,##0.00")
        _cell(ws2, f"E{r}", f"=({avg_input_tokens}*C{r}+{avg_output_tokens}*D{r})/1000000",
              BLACK_FONT, bg, "$#,##0.0000")
        _cell(ws2, f"F{r}", f"=E{r}*48", BLACK_FONT_B, bg, "$#,##0.00")

    nr = 4 + len(MODEL_PRICING) + 1
    ws2.merge_cells(f"B{nr}:F{nr}")
    _cell(ws2, f"B{nr}",
          f"Assumes avg {avg_input_tokens} input tokens and {avg_output_tokens} output tokens per request (from actual run).",
          NOTE_FONT, align=LEFT, border=Border())

    # ── Sheet 3: Actual Run Data ─────────────────────────────────────
    ws3 = wb.create_sheet("Run Data")
    ws3.sheet_properties.tabColor = "28A745"
    ws3.column_dimensions["A"].width = 4
    ws3.column_dimensions["B"].width = 32
    ws3.column_dimensions["C"].width = 22

    for c in range(1, 4):
        ws3.cell(row=1, column=c).fill = DARK_BG
    ws3.merge_cells("B1:C1")
    _cell(ws3, "B1", f"Actual Run: {run_id}", WHITE_FONT_LG,
          align=Alignment(horizontal="left", vertical="center"))
    ws3.row_dimensions[1].height = 36

    actual_rows = [
        ("Model", model_name, None),
        ("Examples Generated", n_examples, "#,##0"),
        ("Run Epochs", run_epochs, "#,##0"),
        ("", "", None),
        ("Total Input Tokens", cost_info.get("total_input_tokens", 0), "#,##0"),
        ("Total Output Tokens", cost_info.get("total_output_tokens", 0), "#,##0"),
        ("Total API Cost", total_cost, "$#,##0.0000"),
        ("Cost per Example", cost_per_example, "$#,##0.0000"),
        ("", "", None),
        ("Teacher Data Generation", f"{api_total_sec:.1f} sec", None),
        ("Baseline Inference", f"{baseline_sec:.1f} sec", None),
        ("LoRA Fine-Tuning", f"{train_sec:.1f} sec", None),
        ("Distilled Inference", f"{distill_sec:.1f} sec", None),
        ("Total Pipeline Time", f"{run_results.get('total_elapsed_seconds', 0):.0f} sec", None),
    ]

    for i, (label, val, fmt) in enumerate(actual_rows):
        r = 3 + i
        ws3.row_dimensions[r].height = 22
        if not label:
            continue
        bg = LIGHT_BG if i < 8 else ALT_BG
        _cell(ws3, f"B{r}", label, BLACK_FONT, bg, align=LEFT)
        _cell(ws3, f"C{r}", val, BLACK_FONT_B, bg, fmt=fmt)

    # Save
    if output is None:
        output = str(run_path / "cost_impact_analysis.xlsx")
    wb.save(output)
    print(f"\n📊 Cost impact analysis saved to {output}")
    return output


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate cost impact analysis from a pipeline run")
    parser.add_argument("run_dir", help="Path to a completed pipeline run directory")
    parser.add_argument("--output", "-o", default=None, help="Output xlsx path (default: <run_dir>/cost_impact_analysis.xlsx)")
    args = parser.parse_args()
    generate_cost_estimate(run_dir=args.run_dir, output=args.output)
